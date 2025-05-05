from datetime import date, datetime, time
from django.shortcuts import render
from django.utils.timezone import now 
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework import status
from attendance.models import Attendance, AttendanceConfig
from .serializers import AttendanceSerializer, AttendanceConfigSerializer

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
class AttendanceHistoryView(APIView):

    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.user
       
        if user.is_authenticated:
            if user.is_superuser:
                attendances = Attendance.objects.all().order_by('-date')
            else:
                attendances = Attendance.objects.filter(employeeId=user.employee).order_by('-date')
        serializer = AttendanceSerializer(attendances, many=True)
        if user.is_superuser:
            data = []
            for att in attendances:
                data.append({
                    "id": att.id,
                    "date": att.date,
                    "check_in": att.check_in,
                    "check_out": att.check_out,
                    "status": att.status,
                    "created_at": att.created_at,
                    "updated_at": att.updated_at,
                    "employee": {
                        "employee_code": att.employeeId.employee_code,
                        "employeeName": att.employeeId.full_name(),
                        "department": att.employeeId.department.name if att.employeeId.department else None,
                    }
                })
            return Response(data)
        else:
            return Response(serializer.data)
class AttendanceConfigView(APIView):
    permission_classes = [IsAdminUser]
    def get(self, request):
        configTime = AttendanceConfig.objects.order_by("-created_at").first()
        if configTime:
            serializer = AttendanceConfigSerializer(configTime)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response({"message": "Chưa có cấu hình thời gian."})
    def post(self, request):
        serializer = AttendanceConfigSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status= status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def put(self, request):
        configTime = AttendanceConfig.objects.order_by("-created_at").first()
        if not configTime:
            return Response({"message": "Chưa có cấu hình thời gian."}, status=status.HTTP_404_NOT_FOUND)
        serializer = AttendanceConfigSerializer(configTime, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class ExportAttendanceExcel(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        param_date = request.query_params.get('date')
        from_date = request.query_params.get('fromDate')
        to_date = request.query_params.get('toDate')
        try:
            if param_date:
                list_date = [
                    datetime.strptime(dates.strip(), '%d/%m/%Y').date()
                    for dates in param_date.split(',')
                ]
                attendances = Attendance.objects.filter(date__in=list_date).order_by('-date')
            elif from_date and to_date:
                from_date = datetime.strptime(from_date.strip(), '%d/%m/%Y').date()
                to_date = datetime.strptime(to_date.strip(), '%d/%m/%Y').date()
                attendances = Attendance.objects.filter(date__range=(from_date, to_date)).order_by('-date')
            elif from_date:
                from_date = datetime.strptime(from_date.strip(), '%d/%m/%Y').date()
                attendances = Attendance.objects.filter(date__gte=from_date).order_by('-date')
            elif to_date:
                to_date = datetime.strptime(to_date.strip(), '%d/%m/%Y').date()
                attendances = Attendance.objects.filter(date__lte=to_date).order_by('-date')
            else:
                attendances = Attendance.objects.all().order_by('-date')
        except ValueError:
            return Response({
                "error": "Sai định dạng ngày. Định dạng hợp lệ: DD/MM/YYYY hoặc danh sách cách nhau bởi dấu phẩy (,)."
            })
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Attendance Report'

        worksheet.merge_cells('A1:G2')
        title_cell = worksheet['A1']
        title_cell.value = "BÁO CÁO CHẤM CÔNG"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        headers = [
            "Mã nhân viên", "Họ và tên", "Phòng ban", "Ngày", "Giờ vào", "Giờ ra", "Trạng thái"
        ]
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True, color="000000")
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for row_num, att in enumerate(attendances, start=4):
            fill = fill_even if row_num % 2 == 0 else fill_odd
            values = [
                att.employeeId.employee_code,
                att.employeeId.full_name(),
                att.employeeId.department.name if att.employeeId.department else '',
                att.date.strftime("%d/%m/%Y"),
                att.check_in.strftime("%H:%M:%S") if att.check_in else '',
                att.check_out.strftime("%H:%M:%S") if att.check_out else '',
                att.status,
            ]
            for col_num, value in enumerate(values, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.fill = fill
                cell.border = thin_border
                if col_num in [4, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            adjusted_width = max_length + 2
            worksheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        worksheet.freeze_panes = 'A4' 

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="attendance_report.xlsx"'
        workbook.save(response)
        return response
